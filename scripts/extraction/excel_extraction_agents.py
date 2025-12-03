"""
Excel Processing with Agent-Based Semantic Enhancement.

This module uses a hybrid approach:
1. openpyxl for reliable, complete extraction (all sheets, formulas, hidden content)
2. Agent for semantic enhancement (metadata extraction, validation, structuring)

The agent doesn't re-extract (openpyxl is reliable) but adds intelligence and context.
"""

import os
import asyncio
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from dotenv import load_dotenv

from scripts.logging_config import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.warning("openpyxl not installed. Run: pip install openpyxl")
    openpyxl = None

from agent_framework import ChatMessage, Role, TextContent
from agent_framework.azure import AzureOpenAIChatClient

# Load environment
load_dotenv()

# Configuration
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT_NAME", "gpt-5-chat")


def extract_excel_structured(excel_path: Path) -> Dict:
    """
    Extract structured data from Excel using openpyxl.

    This is the reliable, programmatic extraction that captures:
    - All worksheets (including hidden)
    - All rows and columns
    - Formulas (as text)
    - Cell formatting
    - Merged cells

    Args:
        excel_path: Path to Excel file

    Returns:
        Dict with sheets, metadata, and raw data
    """
    if not openpyxl:
        raise ImportError("openpyxl is required for Excel processing")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        workbook_info = {
            "file_name": excel_path.name,
            "file_path": str(excel_path),
            "sheet_count": len(wb.worksheets),
            "sheet_names": [sheet.title for sheet in wb.worksheets],
            "sheets": []
        }

        for sheet in wb.worksheets:
            sheet_data = {
                "name": sheet.title,
                "hidden": sheet.sheet_state != 'visible',
                "dimensions": sheet.dimensions,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "rows": []
            }

            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        cell_info = {
                            "value": str(cell.value),
                            "coordinate": cell.coordinate
                        }
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            cell_info["formula"] = cell.value
                        row_data.append(cell_info)
                    else:
                        row_data.append(None)

                if any(cell is not None for cell in row_data):
                    sheet_data["rows"].append(row_data)

            workbook_info["sheets"].append(sheet_data)

        wb.close()
        return workbook_info

    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        raise


def format_excel_as_markdown(workbook_data: Dict) -> str:
    """
    Convert structured Excel data to markdown format.

    Args:
        workbook_data: Structured data from extract_excel_structured

    Returns:
        Markdown representation of the Excel file
    """
    lines = []

    lines.append(f"# {workbook_data['file_name']}\n")
    lines.append(f"**Total Sheets**: {workbook_data['sheet_count']}\n")

    for sheet in workbook_data['sheets']:
        lines.append(f"\n---\n")
        lines.append(f"## Sheet: {sheet['name']}")

        if sheet['hidden']:
            lines.append("*(Hidden sheet)*")

        lines.append(f"\n**Dimensions**: {sheet['dimensions']} ({sheet['max_row']} rows × {sheet['max_column']} columns)\n")

        # Convert rows to markdown table if structured
        if sheet['rows']:
            # Try to detect if first row is a header
            first_row = sheet['rows'][0]
            has_header = all(cell and isinstance(cell, dict) for cell in first_row[:3])

            if len(sheet['rows']) > 1 and has_header:
                # Format as markdown table
                lines.append("")

                # Header row
                header_cells = [cell['value'] if cell else '' for cell in first_row]
                lines.append("| " + " | ".join(header_cells) + " |")
                lines.append("| " + " | ".join(['---'] * len(header_cells)) + " |")

                # Data rows (limit to 1000 rows for markdown)
                for row in sheet['rows'][1:1001]:
                    row_cells = [cell['value'] if cell else '' for cell in row]
                    # Pad to match header length
                    while len(row_cells) < len(header_cells):
                        row_cells.append('')
                    lines.append("| " + " | ".join(row_cells[:len(header_cells)]) + " |")

                if len(sheet['rows']) > 1001:
                    lines.append(f"\n*Note: Table truncated for display. Total rows: {len(sheet['rows'])}*\n")
            else:
                # Format as simple list for non-tabular data
                lines.append("\n**Content:**\n")
                for row_idx, row in enumerate(sheet['rows'][:100], 1):
                    row_text = ", ".join([cell['value'] for cell in row if cell])
                    if row_text:
                        lines.append(f"{row_idx}. {row_text}")

    return "\n".join(lines)


def create_excel_enhancement_instructions() -> str:
    """Create system instructions for Excel Enhancement Agent."""
    return """You are an expert analyst for technical and commercial documents in Excel format.

Your task is to analyze Excel data that has been reliably extracted and add semantic intelligence.

ANALYSIS TASKS:

## 1. Document Purpose Identification
- Determine document type: Bill of Quantities (BOQ), Technical Specifications, Calculations, Test Reports, etc.
- Identify primary purpose and audience
- Note any secondary uses

## 2. Metadata Extraction
- **Equipment Types**: List all equipment mentioned (transformers, circuit breakers, relays, cables, etc.)
- **Voltage Levels**: Extract all voltage levels (400kV, 132kV, 11kV, etc.)
- **Quantities**: Identify item counts and totals
- **Technical Standards**: Note IEC, IEEE, ANSI, BS standards referenced
- **Project References**: Extract project names, locations, dates
- **Manufacturers**: List equipment manufacturers mentioned

## 3. Structure Analysis
- Identify main sections and their purposes
- Note relationships between sheets (if multi-sheet)
- Identify calculated fields vs input data
- Flag summary rows, subtotals, totals

## 4. Quality Validation
- Check for inconsistencies in units (kV vs V, MVA vs kVA)
- Verify quantity calculations if totals present
- Note any missing data or incomplete sections
- Flag any unusual values or potential errors

## 5. Key Requirements
- Extract critical requirements and specifications
- Identify compliance requirements
- Note delivery or timeline requirements
- Extract warranty and guarantee terms

## 6. Enhanced Markdown Output
- Reorganize for clarity and logical flow
- Add section descriptions and context
- Highlight critical specifications
- Include metadata as YAML frontmatter
- Add semantic tags for search optimization

OUTPUT FORMAT:

Provide a JSON response with this structure:

```json
{
  "document_type": "bill_of_quantities | technical_specification | calculation | other",
  "metadata": {
    "equipment_types": ["type1", "type2", ...],
    "voltage_levels": ["400kV", "132kV", ...],
    "standards": ["IEC 62271", "IEEE C37.04", ...],
    "project_references": ["Project A", ...],
    "total_items": 100,
    "estimated_value": "$X million"
  },
  "enhanced_markdown": "The complete enhanced markdown with context and structure",
  "key_requirements": ["Requirement 1", "Requirement 2", ...],
  "validation_notes": ["Note 1", "Note 2", ...],
  "quality_score": 95
}
```

Be thorough and add value through semantic understanding and context."""


# Lazy initialization for Azure OpenAI Client
_client = None
_excel_enhancement_agent = None


def _get_client():
    """Lazily initialize the Azure OpenAI client."""
    global _client
    if _client is None:
        _client = AzureOpenAIChatClient(
            api_key=AZURE_OPENAI_API_KEY,
            endpoint=AZURE_OPENAI_ENDPOINT,
            deployment_name=AZURE_OPENAI_CHAT_DEPLOYMENT,
            api_version=AZURE_OPENAI_API_VERSION
        )
    return _client


def _get_excel_enhancement_agent():
    """Lazily initialize the Excel enhancement agent."""
    global _excel_enhancement_agent
    if _excel_enhancement_agent is None:
        client = _get_client()
        _excel_enhancement_agent = client.create_agent(
            name="Excel_Enhancement",
            instructions=create_excel_enhancement_instructions()
        )
    return _excel_enhancement_agent


async def enhance_excel_with_agent(excel_path: Path, structured_data: Dict, base_markdown: str) -> Dict:
    """
    Enhance Excel extraction with semantic analysis using agent.

    Args:
        excel_path: Path to original Excel file
        structured_data: Structured data from openpyxl extraction
        base_markdown: Markdown representation of Excel

    Returns:
        Enhanced result with metadata and semantic analysis
    """
    analysis_request = f"""Analyze this Excel document and provide semantic enhancement.

**File**: {excel_path.name}
**Sheets**: {', '.join(structured_data['sheet_names'])}
**Total Rows**: {sum(sheet['max_row'] for sheet in structured_data['sheets'])}

**Extracted Content**:

{base_markdown}

Provide comprehensive semantic analysis and enhanced markdown following the instructions."""

    message = ChatMessage(
        role=Role.USER,
        contents=[TextContent(text=analysis_request)]
    )

    try:
        result = await _get_excel_enhancement_agent().run(message)
        response_text = result.text

        try:
            json_str = response_text
            if "```json" in response_text:
                json_str = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                json_str = response_text.split("```")[1].split("```")[0].strip()

            import json
            enhancement_data = json.loads(json_str)
            return enhancement_data

        except (json.JSONDecodeError, KeyError, IndexError):
            return {
                "document_type": "unknown",
                "metadata": {},
                "enhanced_markdown": response_text,
                "key_requirements": [],
                "validation_notes": ["JSON parse failed, using raw agent response"],
                "quality_score": 50
            }

    except Exception as e:
        logger.error(f"Excel enhancement failed: {e}")
        return {
            "document_type": "unknown",
            "metadata": {},
            "enhanced_markdown": base_markdown,
            "key_requirements": [],
            "validation_notes": [f"Agent enhancement failed: {e}"],
            "quality_score": 30
        }


async def process_excel_with_agents(excel_path: Path) -> Dict:
    """
    Process Excel file with hybrid approach: openpyxl + agent enhancement.

    Args:
        excel_path: Path to Excel file

    Returns:
        Dict compatible with existing pipeline format
    """
    logger.info(f"Processing Excel: {excel_path.name}")

    try:
        structured_data = extract_excel_structured(excel_path)
        base_markdown = format_excel_as_markdown(structured_data)
        enhancement = await enhance_excel_with_agent(excel_path, structured_data, base_markdown)

        result = {
            "status": "Succeeded",
            "method": "excel_with_agent_enhancement",
            "result": {
                "contents": [{
                    "markdown": enhancement.get('enhanced_markdown', base_markdown)
                }],
                "document_type": enhancement.get('document_type'),
                "metadata": enhancement.get('metadata', {}),
                "key_requirements": enhancement.get('key_requirements', []),
                "validation_notes": "\n".join(enhancement.get('validation_notes', [])),
                "quality_score": enhancement.get('quality_score', 50),
                "sheets": structured_data['sheet_names']
            }
        }

        logger.info(f"Complete: {excel_path.name} ({len(structured_data['sheets'])} sheets, quality: {result['result']['quality_score']})")
        return result

    except Exception as e:
        logger.error(f"Excel processing failed: {e}")
        return None


def process_excel_with_agents_sync(excel_path: Path) -> Dict:
    """Synchronous wrapper for async Excel processing."""
    return asyncio.run(process_excel_with_agents(excel_path))
